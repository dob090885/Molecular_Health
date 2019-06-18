import openpyxl

def createSheet(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    if sheetName in workbook.sheetnames:
        print("Sheet Exist")
    else:
        workbook.create_sheet(sheetName)
        workbook.save(file)

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=colnum).value

def writeData(file, sheetName, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)
